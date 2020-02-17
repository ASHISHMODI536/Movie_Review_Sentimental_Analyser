#Procedure for using this script
#This script is special design for retriving data from bookmyshow.com html review pages
#Step 1] open the website and see all reviews
#Step 2] Copy the html code from developer mode in browser to notepad and save Encoding type UTF-8
#Setp 3] And follow script instruction

import re
from bs4 import BeautifulSoup as BS 
import bleach
import openpyxl
import emoji

file_name = input("ENTER TEXT FILE :")
excel_file = input("ENTER EXCEL FILE : ")
file = open(file_name,"r",encoding="utf-8")
location = "C:\\Users\\RAHUL\\Desktop\\abc.xlsx"
caption = []
comment = []

soup = BS(file, "html.parser")
data = soup.findAll("div",{"class":"__reviewer-comment"})
for i in data:
    com = str(i.findChild("span",{"class":"__reviewer-left"}))
    caption.append(bleach.clean(str(i), tags=[], attributes={}, styles=[], strip=True))
    
comments = soup.findAll("div",{"class":"__reviewer-text"})

for i in comments:
    comment.append(bleach.clean(str(i), tags=[], attributes={}, styles=[], strip=True))


comment_emoji = []
# for i in comment:
#     try:
#         print(i)
#     except UnicodeEncodeError:
#         comment_emoji.append(comment.pop(comment.index(i)))    

# >>> for i in a:
# ...     if e.demojize(i) == i:
# ...             print("False")
# ...     else:
# ...             print("True")   

for line in comment:
    for i in line:
        if emoji.demojize(i) != i:
            comment_emoji.append(comment.pop(comment.index(line)))
            break

movie_name = soup.find("h1",{"class":"__name"})
movie_name = bleach.clean(str(movie_name), tags=[], attributes={}, styles=[], strip=True)
movie_name = re.sub("[\n\t]","",movie_name)

movie_language = soup.find("a",{"class":"__language"})
movie_language = bleach.clean(str(movie_language), tags=[], attributes={}, styles=[], strip=True)
movie_language = re.sub("[\n\t]","",movie_language)

movie_genre = soup.findAll("span",{"class":"__genre-tag"})
movie_genre = bleach.clean(str(movie_genre), tags=[], attributes={}, styles=[], strip=True)
movie_genre = re.sub("[\n\t\]\[]","",movie_genre)

movie_release_date = soup.find("span",{"class":"__release-date"})
movie_release_date = bleach.clean(str(movie_release_date), tags=[], attributes={}, styles=[], strip=True)
movie_release_date = re.sub("[\n\t\]\[]","",movie_release_date)





wb = openpyxl.Workbook()
sheet = wb.active
wb.create_sheet("emojis")
sheet2 = wb["emojis"]
sheet.cell(1,1).value = "Movie Name"
sheet.cell(1,2).value =  movie_name
sheet.cell(2,1).value = "Movie Language"
sheet.cell(2,2).value =  movie_language
sheet.cell(3,1).value = "Movie Genre"
sheet.cell(3,2).value =  movie_genre
sheet.cell(4,1).value = "Movie Relese Date"
sheet.cell(4,2).value =  movie_release_date
sheet.cell(5,1).value = "Id"
sheet.cell(5,2).value = "Comment"
sheet2.cell(1,1).value = "Id"
sheet2.cell(1,2).value = "Comment"
def initial_comment_write_index(sheet):
    for i in range(1,sheet.max_row+1):
        if sheet.cell(i,2).value == 'Comment':
            return i
        

for i in range(initial_comment_write_index(sheet),len(comment)):
    cell = sheet.cell(i+1,2)
    try:
        cell.value = comment[i]
        sheet.cell(i+1,1).value = i-4
    except IndexError:
        pass

for i in range(initial_comment_write_index(sheet2),len(comment)):
    cell = sheet2.cell(i+1,2)
    try:
        cell.value = comment_emoji[i]
        sheet2.cell(i+1,1).value = i
    except IndexError:
        pass
     

wb.save(excel_file+".xlsx")
wb.close()



