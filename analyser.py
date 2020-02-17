#
#Movie Review Analyser comment predict and score
#
#Importing  required libraries for the program




import openpyxl
import os
import warnings

from pathlib import Path

from nltk.corpus import *
from nltk.classify import NaiveBayesClassifier
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from nltk.tokenize import word_tokenize

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC


warnings.simplefilter(action='ignore')
count = 0


#modified print function for the program
def style_print(data,type="header"):
    if type == "header":
        print("\n+{:-^58}+".format("-"))
        print("|{: ^58}|".format("[ "+data+" ]"))
        print("+{:-^58}+".format("-"))
    elif type == "subheader":
        print("\n<{:-^30}>\n".format("[ "+data+" ]"))
    elif type == "data":
        print(">>> ",data)


#comment grabbing file calling function
#graber.py is used to store comment in excel file
def grab():
    os.system('python '+os.getcwd()+'\\graber.py')
    

#Loading Excel File and returing the location  
def load_Excel():

    #Function for displaying files in specified path
    def file_list(paths):
        list_file = os.listdir(paths)
        print("{:^15}".format("id"), "{:^20}".format("File Name"), sep="|")
        print("{:^43}".format("-" * 38))
        for i in range(len(list_file)):
            print("{:^15}".format(i + 1), "{:^20}".format(list_file[i]), sep="|")
        try:
            choice = int(input("Enter Choice :  "))
        except:
            load_Excel()
        if choice > len(list_file):
            print("Input Error ")
            load_Excel()
        else:
            Excel_path = str(paths) +"\\"+ str(list_file[choice - 1])
            name = list_file[choice - 1]
        print("You have selected ", name ," file")
        return Excel_path

    #Getting the file path    
    print("{:^25}{:^25}".format(" [1] This Folder "," [2] Custom Folder "))
    opt = input("Enter Option : ")
    if opt == '1':
        location = file_list(os.getcwd())
    elif opt == '2':
        try:
            path = input("Enter path : ")
            path = Path(path)
            location = file_list(path)
        except OSError:
            load_Excel()
    else:
        print("INVALID INPUT ....")
        load_Excel()
    return location




#Loading raw data dictionary and returning array of words 
def preprocessing(data):

    stop_words = set(stopwords.words('english'))
    processed_data = []
    for string in data:
        words = word_tokenize(string)                       # tokenizing comments
        words = [word for word in words if word.isalpha()]  # Checking if character is not alphabet remove it
        words = [w for w in words if not w in stop_words]   # Removing all stops words
        words = [w for w in words if not len(w) == 1]       # removing words of length 1
        processed_data.append(words)
    return processed_data



#loading list of words and returnig...
def analysis(words,excel_location,sheet_name=""):
    excel_book = openpyxl.load_workbook(excel_location)
    sheet = excel_book[sheet_name]
    polarity_result = []        #it stores the result of polarity of every comment
    data = []
    neg = neu = pos = com = value = 0
    string = ''
    sid = SentimentIntensityAnalyzer()   
    for i in range(1,sheet.max_row):
        if sheet.cell(i,2).value == 'Comment':
            initial_row_index = i+1
            break
    for i in range(len(words)):
        string = ''
        for j in words[i]:
            string = string + j + " "
        ana = sid.polarity_scores(string)
        polarity_result= ana.values()
        neg, neu, pos, com = polarity_result


        # POLARITY ANALYSIS

        if neg >= 0.4:
            value = -1
        elif pos >= 0.4:
            value = 1
        elif (neg > pos) and (neg - pos) >= 0.15:
            value = -1
        elif (pos > neg) and (pos - neg) >= 0.15 :
            value = 1

        sheet.cell(initial_row_index,3).value = neg
        sheet.cell(initial_row_index,4).value = neu
        sheet.cell(initial_row_index,5).value = pos
        sheet.cell(initial_row_index,6).value = value
        initial_row_index += 1

        data.append([string,value])
    excel_book.save(excel_location)
    return data


#Defining SVM function which perform SVM algoritms

def svm(X_train, X_test, y_train, y_test):
    global count
    count_vect = CountVectorizer()
    try:
        X_train_count = count_vect.fit_transform(X_train)
        X_test_count = count_vect.transform(X_test)
    except ValueError:
        print("File has nothing related to perform !")
        main()
    # SVM ------------
    #running svm algorithm
    try:
        svc = SVC(kernel='rbf', gamma="auto")
        svc.fit(X_train_count, y_train)
        svm_return = svc.predict(X_test_count).mean()
        #print(svm_return)
        if svm_return == 1.0:
            style_print("The movie is positive","data")
        else:
            style_print("The movie is negative","data")
            
    except:
##        style_print("Including False Satatement","data")

        rescue_comment_pos = "This movie very Good Excellent"
        rescue_comment_neg = "This movie very Bad Worst"

        if 1 not in y_train:
            X_train.append(rescue_comment_pos)
            y_train.append(1)
        elif -1 not in y_train:
            X_train.append(rescue_comment_neg)
            y_train.append(-1)

        if 1 not in y_test:
            X_test.append(rescue_comment_pos)
            y_test.append(1)
        elif -1 not in y_test:
            X_test.append(rescue_comment_neg)
            y_test.append(-1)

        svm(X_train, X_test, y_train, y_test)
        
#navie bayes function for getting the sentimental result of the particular comment
def navie_bayes(data):

    # Extract features from the input list of words
    def extract_features(words):
        dicto = dict([(word, True) for word in words])
        return dict([(word, True) for word in words])

    # Load the reviews from the corpus
    fileids_pos = movie_reviews.fileids('pos')
    fileids_neg = movie_reviews.fileids('neg')

    # Extract the features from the reviews
    features_pos = [(extract_features(movie_reviews.words(fileids=[f])), 'Positive') for f in fileids_pos]
    features_neg = [(extract_features(movie_reviews.words(fileids=[f])), 'Negative') for f in fileids_neg]

    # Define the train and test split (80% and 20%)
    threshold = 0.8
    num_pos = int(threshold * len(features_pos))
    num_neg = int(threshold * len(features_neg))

    # Create training and training datasets
    features_train = features_pos[:num_pos] + features_neg[:num_neg]
    features_test = features_pos[num_pos:] + features_neg[num_neg:]


    # Train a Naive Bayes classifier
    classifier = NaiveBayesClassifier.train(features_train)

    # Test input movie reviews
    input_reviews = data
    count_neg = 0
    count_pos = 0
    neg_vec = []
    pos_vec= []
    for review in input_reviews:
        probabilities = classifier.prob_classify(extract_features(review.split()))

        # Pick the maximum value
        predicted_sentiment = probabilities.max()

        # Print outputs
        if (predicted_sentiment == "Negative"):
            count_neg += 1
            neg_vec.append(round(probabilities.prob(predicted_sentiment), 2))
        else:
            count_pos += 1
            pos_vec.append(round(probabilities.prob(predicted_sentiment), 2))
		
    pos = neg = 0
    for i in pos_vec:
        pos = pos + i
    for i in neg_vec:
        neg = neg + i
    total = neg + pos
    return [(pos/total)*100,(neg/total)*100]


#function for sequencing the flow of program
def start():


    location = ''
    print("{:^25}{:^25}".format(" [1] Load Excel File "," [2] GO BACK "))
    opt = input("Enter Option : ")
    if opt == '1':
        print("Loading Excel file..... ")
        location = load_Excel()
        print(location)
        try:
            wb = openpyxl.load_workbook(location)
        except:
            print("File Not Found!")
            start()
        try:
            sheet1 = wb.active
            sheet2 = wb['emojis']
        except KeyError:
            print("File structure is not supported by program!")
            print("File has no sheet name --> emojis")
            main()
        except UnboundLocalError:
            pass

        #Defining list that will store comments data
        comment_sheet_1 = []
        comment_sheet_2 = []
        value = 0
        
        #---------------------------------------------------------------------#
        # Data retrive is used for extracting comment from excel sheet 
        def data_retrive(sheet):
            comment_sheet = []
            for i in range(1,sheet.max_row):
                if sheet.cell(i,2).value == 'Comment':
                    start_row_index = i
                    sheet.cell(i,3).value = "Negative"
                    sheet.cell(i,4).value = "Neutral"
                    sheet.cell(i,5).value = "Positive"
                    sheet.cell(i,6).value = "value"
                    wb.save(location)
                    break
            for i in range(start_row_index+1,sheet.max_row+1):
                value = sheet.cell(i,2).value
                if value != None:
                    comment_sheet.append(value)
            return comment_sheet

        
        style_print("MOVIE NAME : " + sheet1.cell(1,2).value)

        #stored data in list
        comment_sheet_1 = data_retrive(sheet1)
        comment_sheet_2 = data_retrive(sheet2)

        #data of list is being preprocessed
        #preprocessed trims the comment which removed unwanted characters for example puntuation marks
        comment_data_sheet1 = preprocessing(comment_sheet_1)
        comment_data_sheet2 = preprocessing(comment_sheet_2)
        comment_value_sheet1 = analysis(comment_data_sheet1,location,sheet_name="Sheet")
        comment_value_sheet2 = analysis(comment_data_sheet2,location,sheet_name="emojis")

        #Spliting data into test and train set
        text_train, text_test = train_test_split(comment_value_sheet1, test_size=0.3, random_state= 1)
        emoji_train, emoji_test = train_test_split(comment_value_sheet2, test_size=0.3, random_state=1)

        X_text_train = X_text_test = y_text_train = y_text_test = []
        y_emoji_train = X_emoji_train = y_emoji_test = X_emoji_test = []

        X_text_train = [w[0] for w in text_train ]
        y_text_train = [w[1] for w in text_train ]
        X_text_test = [w[0] for w in text_test]
        y_text_test = [w[1] for w in text_test]
        X_emoji_train = [w[0] for w in emoji_train]
        y_emoji_train = [w[1] for w in emoji_train]
        X_emoji_test = [w[0] for w in emoji_test]
        y_emoji_test = [w[1] for w in emoji_test]

        # SVM ---------------------------------------------
        style_print("SVM - SVC ALGORITHM")
        style_print("TEXT","subheader")
        svm(X_text_train,X_text_test,y_text_train,y_text_test)
        style_print("Emoji","subheader")
        svm(X_emoji_train,X_emoji_test,y_emoji_train,y_emoji_test)

        # NAVIE BAYES -------------------------------------
        style_print("NB-TRAIN")
        
        #navie_bayes_review function for calculating, aggregating and print the comment result
        def navie_bayes_review(positive,negative,total,type):         
            total_pos = total_neg = 0
            for i in positive:
                total_pos = total_pos + i
            for i in negative:
                total_neg = total_neg + i
            total = total_pos +  total_neg
            total_pos = (total_pos/total)*100
            style_print(type,"subheader")
            style_print("Positive Percentage : "+str(total_pos),"data")
            style_print("Negative Percentage : "+str((total_neg/total)*100),"data")
            if total_pos > 85:
                style_print("Excellent","data")
            elif total_pos > 75:
                style_print("Very Good","data")
            elif total_pos > 65:
                style_print("Good","data")
            elif total_pos > 50:
                style_print("Average","data")
            else:
                style_print("worst","data")
        #---------TEXT-------
        positive = []
        negative = []
        #stores numeric values of comments which expresses the sentiment of comment
        nb_train = navie_bayes(X_text_train)
        nb_test = navie_bayes(X_text_test)
        total = nb_train + nb_test
        positive = [total[i] for i in range(0,len(total),2)]
        negative = [total[i] for i in range(1,len(total),2)]
        #processing text
        navie_bayes_review(positive,negative,total,"TEXT")
        #---------EMOJI-------
        positive = []
        negative = []
        nb_emoji_train = navie_bayes(X_emoji_train)
        nb_emoji_test = navie_bayes(X_emoji_test)
        total = nb_emoji_train + nb_emoji_test
        positive = [total[i] for i in range(0,len(total),2)]
        negative = [total[i] for i in range(1,len(total),2)]
        #process emojis
        navie_bayes_review(positive,negative,total,"EMOJI")
                
    elif opt == '2':
        main()
    else:
        print("INVALID INPUT ....")
        start()        


#Starting of program
def _main():
    print("{:*^100}".format(" Wellcome to movie review analyser "))
    print("{:^30}{:^30}{:^30}".format(" [1] Start "," [2] Grab Data "," [3] Exit "))
    opt = input("Enter Option : ")
    if opt == '1':
        start()
    elif opt == '2':
        grab()
        _main()
    elif opt == '3':
        exit()
    else :
        print("INVALID INPUT ....")
        _main()

#starting the program
_main()


